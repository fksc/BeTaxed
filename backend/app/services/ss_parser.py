"""Parse Segurança Social Vínculos + Contratos extracts (xlsx or csv, DEV-831)."""

from __future__ import annotations

import csv
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from app.security.crypto import normalize_niss
from app.services.ss_headers import (
    CONTRATO_FIELDS,
    CONTRATO_REQUIRED,
    VINCULO_FIELDS,
    VINCULO_REQUIRED,
    fold_header,
    is_ignored_header,
    parse_export_label,
)

SheetKind = Literal["vinculos", "contratos"]
FileKind = Literal["COMBINED_XLSX", "VINCULOS", "CONTRATOS"]


class SsParseError(Exception):
    """File cannot be parsed into official Vínculos + Contratos rows."""

    def __init__(self, message: str) -> None:
        super().__init__(message)
        self.message = message


@dataclass(frozen=True)
class SsParseWarning:
    code: str
    detail: str


@dataclass
class ParsedVinculo:
    source_row: int
    niss: str
    nome: str | None = None
    dob: date | None = None
    vinculo_raw: str | None = None
    communicated_on: date | None = None
    started_on: date | None = None
    ended_on: date | None = None
    rate_from: date | None = None
    rate_to: date | None = None
    taxa_pct: Decimal | None = None
    workplace_ss_label: str | None = None
    leftover: dict[str, Any] = field(default_factory=dict)


@dataclass
class ParsedContrato:
    source_row: int
    niss: str
    nome: str | None = None
    modality_raw: str | None = None
    work_mode_raw: str | None = None
    contract_started_on: date | None = None
    contract_ended_on: date | None = None
    profession_raw: str | None = None
    percent_work: Decimal | None = None
    hours_work: Decimal | None = None
    days_work: Decimal | None = None
    motivo_raw: str | None = None
    rendimento_from: date | None = None
    rendimento_to: date | None = None
    base_salary: Decimal | None = None
    leftover: dict[str, Any] = field(default_factory=dict)


@dataclass
class ParsedWorkbook:
    kind: FileKind
    vinculos: list[ParsedVinculo]
    contratos: list[ParsedContrato]
    employer_niss: str | None
    export_label: str | None


@dataclass
class ParsedSsExport:
    vinculos: list[ParsedVinculo]
    contratos: list[ParsedContrato]
    file_kinds: list[FileKind]
    employer_niss: str | None
    export_label: str | None
    warnings: list[SsParseWarning]


@dataclass(frozen=True)
class SsSourceFile:
    filename: str
    content: bytes


def current_contratos(rows: list[ParsedContrato]) -> list[ParsedContrato]:
    """Current pay = open rendimento period, never the first sheet row."""
    return [row for row in rows if row.rendimento_to is None]


def parse_ss_files(files: list[SsSourceFile]) -> ParsedSsExport:
    if not files:
        raise SsParseError("At least one SS export file is required.")
    if len(files) > 2:
        raise SsParseError("Expected one combined workbook or two files (xlsx or csv).")

    parsed_files = [parse_ss_workbook(item) for item in files]
    vinculos: list[ParsedVinculo] = []
    contratos: list[ParsedContrato] = []
    kinds: list[FileKind] = []
    employer_niss: str | None = None
    export_label: str | None = None

    for item, parsed in zip(files, parsed_files, strict=True):
        kinds.append(parsed.kind)
        vinculos.extend(parsed.vinculos)
        contratos.extend(parsed.contratos)
        if parsed.employer_niss and employer_niss is None:
            employer_niss = parsed.employer_niss
        if parsed.export_label and (
            export_label is None or "_vinculos_" in parsed.export_label.lower()
        ):
            export_label = parsed.export_label
        if employer_niss is None:
            from_name, label = parse_export_label(item.filename)
            if from_name:
                employer_niss = from_name
                export_label = export_label or label

    if not vinculos:
        raise SsParseError("No Vínculos rows found.")
    if not contratos:
        raise SsParseError("No Contratos rows found.")
    if len(files) == 2:
        kinds_set = set(kinds)
        if kinds_set != {"VINCULOS", "CONTRATOS"}:
            raise SsParseError(
                "Two files must be one Vínculos file and one Contratos file."
            )

    warnings = _join_warnings(vinculos, contratos)
    warnings.extend(_open_period_warnings(contratos))
    return ParsedSsExport(
        vinculos=vinculos,
        contratos=contratos,
        file_kinds=kinds,
        employer_niss=employer_niss,
        export_label=export_label,
        warnings=warnings,
    )


def parse_ss_workbook(source: SsSourceFile) -> ParsedWorkbook:
    if source.content.startswith(b"PK"):
        try:
            workbook = load_workbook(BytesIO(source.content), data_only=False)
        except Exception as exc:
            raise SsParseError(
                f"{source.filename} could not be opened as xlsx."
            ) from exc
        try:
            classified = _classify_sheets(source.filename, list(workbook.worksheets))
            return _parsed_workbook(source, classified)
        finally:
            workbook.close()
    classified = _classify_csv(source)
    return _parsed_workbook(source, classified)


def _parsed_workbook(
    source: SsSourceFile, classified: dict[SheetKind, Any]
) -> ParsedWorkbook:
    vinculos: list[ParsedVinculo] = []
    contratos: list[ParsedContrato] = []
    if "vinculos" in classified:
        vinculos = _parse_vinculo_sheet(classified["vinculos"])
    if "contratos" in classified:
        contratos = _parse_contrato_sheet(classified["contratos"])

    if "vinculos" in classified and "contratos" in classified:
        kind: FileKind = "COMBINED_XLSX"
    elif "vinculos" in classified:
        kind = "VINCULOS"
    else:
        kind = "CONTRATOS"

    employer_niss, export_label = parse_export_label(source.filename)
    return ParsedWorkbook(
        kind=kind,
        vinculos=vinculos,
        contratos=contratos,
        employer_niss=employer_niss,
        export_label=export_label,
    )


def _classify_csv(source: SsSourceFile) -> dict[SheetKind, Any]:
    rows = _csv_rows(source)
    sheet = _RowSheet("csv", rows)
    return _classify_sheets(source.filename, [sheet])


def _classify_sheets(filename: str, sheets: list[Any]) -> dict[SheetKind, Any]:
    classified: dict[SheetKind, Any] = {}
    for sheet in sheets:
        kind = _sheet_kind(sheet.title) or _sheet_kind_from_headers(sheet)
        if kind is None:
            continue
        if kind in classified:
            raise SsParseError(f"{filename} has more than one {kind} sheet.")
        classified[kind] = sheet
    if not classified:
        raise SsParseError(
            f"{filename} has no Vínculos or Contratos sheet (xlsx or csv)."
        )
    return classified


class _RowSheet:
    """Minimal sheet so CSV rows reuse the xlsx header/row parsers."""

    def __init__(self, title: str, rows: list[list[Any]]) -> None:
        self.title = title
        self._rows = rows

    def iter_rows(self, values_only: bool = True):
        _ = values_only
        for row in self._rows:
            yield tuple(row)


def _csv_rows(source: SsSourceFile) -> list[list[Any]]:
    text = _decode_csv(source)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t")
        reader = csv.reader(StringIO(text), dialect)
    except csv.Error:
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","
        reader = csv.reader(StringIO(text), delimiter=delimiter)
    rows = [list(row) for row in reader]
    if not rows:
        raise SsParseError(f"{source.filename} is empty.")
    return rows


def _decode_csv(source: SsSourceFile) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            text = source.content.decode(encoding)
        except UnicodeDecodeError:
            continue
        if "\x00" in text[:4096]:
            raise SsParseError(
                f"{source.filename} is not an xlsx workbook or csv extract."
            )
        return text
    raise SsParseError(f"{source.filename} could not be read as csv text.")


def _sheet_kind(title: str) -> SheetKind | None:
    folded = fold_header(title)
    if "vinculo" in folded:
        return "vinculos"
    if "contrato" in folded:
        return "contratos"
    return None


def _sheet_kind_from_headers(sheet: Any) -> SheetKind | None:
    header = _header_row(sheet)
    if header is None:
        return None
    names = {fold_header(str(cell)) for _, cell in header if cell is not None}
    if "local de trabalho" in names:
        return "vinculos"
    if "modalidade contrato" in names:
        return "contratos"
    return None


def _header_row(sheet: Any) -> list[tuple[int, Any]] | None:
    _, row = _first_nonempty_row(sheet)
    if row is None:
        return None
    return list(enumerate(row))


def _first_nonempty_row(sheet: Any) -> tuple[int | None, list[Any] | None]:
    for excel_row, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if any(cell is not None and str(cell).strip() for cell in row):
            return excel_row, list(row)
    return None, None


def _parse_vinculo_sheet(sheet: Any) -> list[ParsedVinculo]:
    mapped, leftover_cols, missing = _map_columns(
        sheet, VINCULO_FIELDS, VINCULO_REQUIRED
    )
    if missing:
        raise SsParseError(
            "Vínculos sheet is missing required headers: " + ", ".join(missing)
        )
    rows: list[ParsedVinculo] = []
    for source_row, values in _iter_data_rows(sheet):
        fields, leftover = _row_values(values, mapped, leftover_cols)
        niss = _optional_niss(fields.get("niss"))
        if niss is None:
            continue
        rows.append(
            ParsedVinculo(
                source_row=source_row,
                niss=niss,
                nome=_optional_str(fields.get("nome")),
                dob=_optional_date(fields.get("dob")),
                vinculo_raw=_optional_str(fields.get("vinculo_raw")),
                communicated_on=_optional_date(fields.get("communicated_on")),
                started_on=_optional_date(fields.get("started_on")),
                ended_on=_optional_date(fields.get("ended_on")),
                rate_from=_optional_date(fields.get("rate_from")),
                rate_to=_optional_date(fields.get("rate_to")),
                taxa_pct=_optional_decimal(fields.get("taxa_pct")),
                workplace_ss_label=_optional_str(fields.get("workplace_ss_label")),
                leftover=leftover,
            )
        )
    if not rows:
        raise SsParseError("Vínculos sheet has no data rows.")
    return rows


def _parse_contrato_sheet(sheet: Any) -> list[ParsedContrato]:
    mapped, leftover_cols, missing = _map_columns(
        sheet, CONTRATO_FIELDS, CONTRATO_REQUIRED
    )
    if missing:
        raise SsParseError(
            "Contratos sheet is missing required headers: " + ", ".join(missing)
        )
    rows: list[ParsedContrato] = []
    for source_row, values in _iter_data_rows(sheet):
        fields, leftover = _row_values(values, mapped, leftover_cols)
        niss = _optional_niss(fields.get("niss"))
        if niss is None:
            continue
        rows.append(
            ParsedContrato(
                source_row=source_row,
                niss=niss,
                nome=_optional_str(fields.get("nome")),
                modality_raw=_optional_str(fields.get("modality_raw")),
                work_mode_raw=_optional_str(fields.get("work_mode_raw")),
                contract_started_on=_optional_date(fields.get("contract_started_on")),
                contract_ended_on=_optional_date(fields.get("contract_ended_on")),
                profession_raw=_optional_str(fields.get("profession_raw")),
                percent_work=_optional_decimal(fields.get("percent_work")),
                hours_work=_optional_decimal(fields.get("hours_work")),
                days_work=_optional_decimal(fields.get("days_work")),
                motivo_raw=_optional_str(fields.get("motivo_raw")),
                rendimento_from=_optional_date(fields.get("rendimento_from")),
                rendimento_to=_optional_date(fields.get("rendimento_to")),
                base_salary=_optional_decimal(fields.get("base_salary")),
                leftover=leftover,
            )
        )
    if not rows:
        raise SsParseError("Contratos sheet has no data rows.")
    return rows


def _map_columns(
    sheet: Any,
    field_map: dict[str, str],
    required: tuple[str, ...],
) -> tuple[dict[int, str], dict[int, str], list[str]]:
    header = _header_row(sheet)
    if header is None:
        raise SsParseError("Sheet has no header row.")

    mapped: dict[int, str] = {}
    leftover_cols: dict[int, str] = {}
    used_fields: set[str] = set()
    seen_canonical: set[str] = set()

    for idx, cell in header:
        if cell is None or not str(cell).strip():
            continue
        original = str(cell).strip()
        if is_ignored_header(original):
            continue
        canonical = fold_header(original)
        field_name = field_map.get(canonical)
        if field_name is None:
            leftover_cols[idx] = original
            continue
        if field_name in used_fields:
            continue
        mapped[idx] = field_name
        used_fields.add(field_name)
        seen_canonical.add(canonical)

    missing = [name for name in required if name not in seen_canonical]
    return mapped, leftover_cols, missing


def _iter_data_rows(sheet: Any) -> list[tuple[int, list[Any]]]:
    rows: list[tuple[int, list[Any]]] = []
    header_row, _ = _first_nonempty_row(sheet)
    if header_row is None:
        return rows
    for excel_row, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if excel_row <= header_row:
            continue
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        rows.append((excel_row, list(row)))
    return rows


def _row_values(
    values: list[Any],
    mapped: dict[int, str],
    leftover_cols: dict[int, str],
) -> tuple[dict[str, Any], dict[str, Any]]:
    fields: dict[str, Any] = {}
    leftover: dict[str, Any] = {}
    for idx, field_name in mapped.items():
        fields[field_name] = values[idx] if idx < len(values) else None
    for idx, header in leftover_cols.items():
        raw = values[idx] if idx < len(values) else None
        cleaned = _leftover_value(raw)
        if cleaned is not None:
            leftover[header] = cleaned
    return fields, leftover


def _leftover_value(raw: Any) -> Any:
    if raw is None or raw == "":
        return None
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if isinstance(raw, date):
        return raw.isoformat()
    if isinstance(raw, Decimal):
        return str(raw)
    return raw


def _optional_str(raw: Any) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    return text or None


def _optional_niss(raw: Any) -> str | None:
    if raw is None or raw == "":
        return None
    digits = normalize_niss(str(raw))
    return digits or None


def _optional_decimal(raw: Any) -> Decimal | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, Decimal):
        return raw
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return Decimal(str(raw))
    text = str(raw).strip().replace(" ", "").replace("%", "")
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _optional_date(raw: Any) -> date | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        try:
            converted = from_excel(raw)
        except Exception:
            # Excel 1900-system serial as days since 1899-12-30.
            converted = datetime(1899, 12, 30) + timedelta(days=float(raw))
        if isinstance(converted, datetime):
            return converted.date()
        if isinstance(converted, date):
            return converted
    if isinstance(raw, str):
        text = raw.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def _join_warnings(
    vinculos: list[ParsedVinculo],
    contratos: list[ParsedContrato],
) -> list[SsParseWarning]:
    vinculo_niss = {row.niss for row in vinculos}
    contrato_niss = {row.niss for row in contratos}
    warnings: list[SsParseWarning] = []
    orphan_contratos = contrato_niss - vinculo_niss
    orphan_vinculos = vinculo_niss - contrato_niss
    if orphan_contratos:
        warnings.append(
            SsParseWarning(
                code="ORPHAN_CONTRATO",
                detail=f"{len(orphan_contratos)} contrato NISS with no vínculo",
            )
        )
    if orphan_vinculos:
        warnings.append(
            SsParseWarning(
                code="ORPHAN_VINCULO",
                detail=f"{len(orphan_vinculos)} vínculo NISS with no contrato",
            )
        )
    return warnings


def _open_period_warnings(contratos: list[ParsedContrato]) -> list[SsParseWarning]:
    open_by_niss: dict[str, list[int]] = defaultdict(list)
    for row in current_contratos(contratos):
        open_by_niss[row.niss].append(row.source_row)
    multiples = {niss: rows for niss, rows in open_by_niss.items() if len(rows) > 1}
    if not multiples:
        return []
    return [
        SsParseWarning(
            code="MULTIPLE_OPEN_RENDIMENTO",
            detail=f"{len(multiples)} people have more than one open rendimento period",
        )
    ]
